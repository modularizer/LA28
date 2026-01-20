"""
Export functions for CSV, JSON, and Excel output.
"""

import csv
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlmodel import select
from sqlalchemy.orm import selectinload

from la28.database import Database
from la28.models import Zone, Venue, Sport, Session, Event


def _dt_to_iso(dt: datetime) -> str:
    """Convert datetime to ISO string for JSON export."""
    return dt.isoformat() if dt else None


def export_sessions_json(db: Database, path: str | Path, indent: int = 2) -> int:
    """Export all sessions with events to JSON."""
    with db.session() as s:
        sessions = s.exec(
            select(Session)
            .options(
                selectinload(Session.events),
                selectinload(Session.sport_rel),
                selectinload(Session.venue_rel),
            )
            .order_by(Session.starts_at)
        ).all()

        data = []
        for sess in sessions:
            d = {
                "code": sess.code,
                "day": sess.day,
                "sport": sess.sport,
                "venue": sess.venue,
                "zone": sess.venue_rel.zone if sess.venue_rel else None,
                "type": sess.type,
                "startsAt": _dt_to_iso(sess.starts_at),
                "endsAt": _dt_to_iso(sess.ends_at),
                "timezone": sess.timezone,
                "durationMinutes": sess.duration_minutes,
                "ticketed": sess.ticketed,
                "inOKC": sess.venue_rel.in_okc if sess.venue_rel else False,
                "sessionNumber": sess.session_number,
                "totalSessions": sess.total_sessions,
                "sportSessionNumber": sess.sport_session_number,
                "totalSportSessions": sess.total_sport_sessions,
                "events": [
                    {
                        "sex": e.sex,
                        "description": e.description,
                        "type": e.type,
                        "order": e.order_in_session,
                        "total": e.total_in_session,
                    }
                    for e in sorted(sess.events, key=lambda x: x.order_in_session)
                ],
            }
            data.append(d)

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=indent)

    return len(data)


def export_events_json(db: Database, path: str | Path, indent: int = 2) -> int:
    """Export all events with session info to JSON."""
    with db.session() as s:
        events = s.exec(
            select(Event)
            .options(
                selectinload(Event.session_rel).selectinload(Session.venue_rel),
            )
            .order_by(Event.event_number)
        ).all()

        data = []
        for e in events:
            sess = e.session_rel
            d = {
                "code": e.code,
                "sport": sess.sport if sess else None,
                "venue": sess.venue if sess else None,
                "zone": sess.venue_rel.zone if sess and sess.venue_rel else None,
                "day": sess.day if sess else None,
                "startsAt": _dt_to_iso(sess.starts_at) if sess else None,
                "sex": e.sex,
                "description": e.description,
                "type": e.type,
                "order_in_session": e.order_in_session,
                "event_number": e.event_number,
            }
            data.append(d)

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=indent)

    return len(data)


def export_sessions_csv(db: Database, path: str | Path) -> int:
    """Export sessions to CSV."""
    with db.session() as s:
        sessions = s.exec(
            select(Session)
            .options(selectinload(Session.venue_rel), selectinload(Session.events))
            .order_by(Session.starts_at)
        ).all()

        fieldnames = [
            "code", "day", "sport", "venue", "zone", "type",
            "starts_at", "ends_at", "timezone", "duration_minutes",
            "ticketed", "in_okc", "num_events",
        ]

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for sess in sessions:
                writer.writerow({
                    "code": sess.code,
                    "day": sess.day,
                    "sport": sess.sport,
                    "venue": sess.venue,
                    "zone": sess.venue_rel.zone if sess.venue_rel else None,
                    "type": sess.type,
                    "starts_at": _dt_to_iso(sess.starts_at),
                    "ends_at": _dt_to_iso(sess.ends_at),
                    "timezone": sess.timezone,
                    "duration_minutes": sess.duration_minutes,
                    "ticketed": sess.ticketed,
                    "in_okc": sess.venue_rel.in_okc if sess.venue_rel else False,
                    "num_events": len(sess.events),
                })

    return len(sessions)


def export_events_csv(db: Database, path: str | Path) -> int:
    """Export events to CSV with session columns."""
    with db.session() as s:
        events = s.exec(
            select(Event)
            .options(selectinload(Event.session_rel).selectinload(Session.venue_rel))
            .order_by(Event.event_number)
        ).all()

        fieldnames = [
            "code", "day", "sport", "venue", "zone", "starts_at",
            "sex", "description", "type", "order_in_session", "event_number",
        ]

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for e in events:
                sess = e.session_rel
                writer.writerow({
                    "code": e.code,
                    "day": sess.day if sess else None,
                    "sport": sess.sport if sess else None,
                    "venue": sess.venue if sess else None,
                    "zone": sess.venue_rel.zone if sess and sess.venue_rel else None,
                    "starts_at": _dt_to_iso(sess.starts_at) if sess else None,
                    "sex": e.sex,
                    "description": e.description,
                    "type": e.type,
                    "order_in_session": e.order_in_session,
                    "event_number": e.event_number,
                })

    return len(events)


def export_sports_json(db: Database, path: str | Path, indent: int = 2) -> int:
    """Export sports with venues."""
    with db.session() as s:
        sports = s.exec(
            select(Sport).options(selectinload(Sport.venues)).order_by(Sport.sport)
        ).all()

        data = [
            {"sport": sp.sport, "venues": [v.venue for v in sp.venues]}
            for sp in sports
        ]

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=indent)

    return len(data)


def export_sports_csv(db: Database, path: str | Path) -> int:
    """Export sports to CSV."""
    with db.session() as s:
        sports = s.exec(
            select(Sport).options(selectinload(Sport.venues)).order_by(Sport.sport)
        ).all()

        fieldnames = ["sport", "description", "icon", "venues"]

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for sp in sports:
                writer.writerow({
                    "sport": sp.sport,
                    "description": sp.description,
                    "icon": sp.icon,
                    "venues": "; ".join(v.venue for v in sp.venues),
                })

    return len(sports)


def export_venues_json(db: Database, path: str | Path, indent: int = 2) -> int:
    """Export venues with zone and sports."""
    with db.session() as s:
        venues = s.exec(
            select(Venue).options(selectinload(Venue.sports)).order_by(Venue.venue)
        ).all()

        data = [
            {
                "venue": v.venue,
                "zone": v.zone,
                "in_okc": v.in_okc,
                "sports": [sp.sport for sp in v.sports],
            }
            for v in venues
        ]

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=indent)

    return len(data)


def export_venues_csv(db: Database, path: str | Path) -> int:
    """Export venues to CSV."""
    with db.session() as s:
        venues = s.exec(
            select(Venue).options(selectinload(Venue.sports)).order_by(Venue.venue)
        ).all()

        fieldnames = ["venue", "zone", "address", "capacity", "latitude", "longitude", "in_okc", "sports"]

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for v in venues:
                writer.writerow({
                    "venue": v.venue,
                    "zone": v.zone,
                    "address": v.address,
                    "capacity": v.capacity,
                    "latitude": v.latitude,
                    "longitude": v.longitude,
                    "in_okc": v.in_okc,
                    "sports": "; ".join(sp.sport for sp in v.sports),
                })

    return len(venues)


def export_zones_json(db: Database, path: str | Path, indent: int = 2) -> int:
    """Export zones with venues."""
    with db.session() as s:
        zones = s.exec(
            select(Zone).options(selectinload(Zone.venues)).order_by(Zone.zone)
        ).all()

        data = [
            {"zone": z.zone, "venues": [v.venue for v in z.venues]}
            for z in zones
        ]

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=indent)

    return len(data)


def export_zones_csv(db: Database, path: str | Path) -> int:
    """Export zones to CSV."""
    with db.session() as s:
        zones = s.exec(
            select(Zone).options(selectinload(Zone.venues)).order_by(Zone.zone)
        ).all()

        fieldnames = ["zone", "description", "venues"]

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for z in zones:
                writer.writerow({
                    "zone": z.zone,
                    "description": z.description,
                    "venues": "; ".join(v.venue for v in z.venues),
                })

    return len(zones)


def export_xlsx(db: Database, path: str | Path) -> dict[str, int]:
    """Export all data to a multi-sheet Excel workbook with proper formatting and links."""
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import NamedStyle, Border, Side

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")
    date_format = "YYYY-MM-DD HH:MM"
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    def style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    def auto_width(ws):
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    def add_borders(ws, row_count):
        for row in ws.iter_rows(min_row=2, max_row=row_count + 1):
            for cell in row:
                cell.border = thin_border

    counts = {}

    # First pass: collect lookup data for links and validation
    with db.session() as s:
        all_sports = {sp.sport: i + 2 for i, sp in enumerate(
            s.exec(select(Sport).order_by(Sport.sport)).all()
        )}
        all_venues = {v.venue: i + 2 for i, v in enumerate(
            s.exec(select(Venue).order_by(Venue.venue)).all()
        )}
        all_zones = {z.zone: i + 2 for i, z in enumerate(
            s.exec(select(Zone).order_by(Zone.zone)).all()
        )}
        all_sessions = {sess.code: i + 2 for i, sess in enumerate(
            s.exec(select(Session).order_by(Session.starts_at)).all()
        )}

    # Sports sheet (create first for validation references)
    ws_sports = wb.create_sheet("Sports")
    with db.session() as s:
        sports = s.exec(
            select(Sport).options(selectinload(Sport.venues)).order_by(Sport.sport)
        ).all()

        ws_sports.append(["sport", "description", "icon", "num_venues"])
        for sp in sports:
            ws_sports.append([sp.sport, sp.description, sp.icon, len(sp.venues)])

        # Format numbers
        for row in range(2, len(sports) + 2):
            ws_sports.cell(row=row, column=4).number_format = '0'

        counts["sports"] = len(sports)
    style_header(ws_sports)
    add_borders(ws_sports, len(sports))
    auto_width(ws_sports)

    # Zones sheet
    ws_zones = wb.create_sheet("Zones")
    with db.session() as s:
        zones = s.exec(
            select(Zone).options(selectinload(Zone.venues)).order_by(Zone.zone)
        ).all()

        ws_zones.append(["zone", "description", "num_venues"])
        for z in zones:
            ws_zones.append([z.zone, z.description, len(z.venues)])

        for row in range(2, len(zones) + 2):
            ws_zones.cell(row=row, column=3).number_format = '0'

        counts["zones"] = len(zones)
    style_header(ws_zones)
    add_borders(ws_zones, len(zones))
    auto_width(ws_zones)

    # Venues sheet
    ws_venues = wb.create_sheet("Venues")
    with db.session() as s:
        venues = s.exec(
            select(Venue).options(selectinload(Venue.sports)).order_by(Venue.venue)
        ).all()

        ws_venues.append(["venue", "zone", "address", "capacity", "latitude", "longitude", "in_okc", "num_sports"])

        for row_idx, v in enumerate(venues, 2):
            ws_venues.cell(row=row_idx, column=1, value=v.venue)

            # Zone as link
            if v.zone and v.zone in all_zones:
                cell = ws_venues.cell(row=row_idx, column=2)
                cell.value = f'=HYPERLINK("#Zones!A{all_zones[v.zone]}", "{v.zone}")'
                cell.font = link_font
            else:
                ws_venues.cell(row=row_idx, column=2, value=v.zone)

            ws_venues.cell(row=row_idx, column=3, value=v.address)
            cap_cell = ws_venues.cell(row=row_idx, column=4, value=v.capacity)
            cap_cell.number_format = '#,##0'
            ws_venues.cell(row=row_idx, column=5, value=v.latitude)
            ws_venues.cell(row=row_idx, column=6, value=v.longitude)
            ws_venues.cell(row=row_idx, column=7, value=v.in_okc)
            ws_venues.cell(row=row_idx, column=8, value=len(v.sports))

        counts["venues"] = len(venues)
    style_header(ws_venues)
    add_borders(ws_venues, len(venues))
    auto_width(ws_venues)

    # Sessions sheet
    ws = wb.active
    ws.title = "Sessions"
    with db.session() as s:
        sessions = s.exec(
            select(Session)
            .options(selectinload(Session.venue_rel), selectinload(Session.events))
            .order_by(Session.starts_at)
        ).all()

        ws.append(["code", "day", "sport", "venue", "zone", "type", "starts_at", "ends_at",
                   "duration_min", "ticketed", "in_okc", "num_events"])

        for row_idx, sess in enumerate(sessions, 2):
            ws.cell(row=row_idx, column=1, value=sess.code)

            day_cell = ws.cell(row=row_idx, column=2, value=sess.day)
            day_cell.number_format = '0'

            # Sport as link
            if sess.sport in all_sports:
                cell = ws.cell(row=row_idx, column=3)
                cell.value = f'=HYPERLINK("#Sports!A{all_sports[sess.sport]}", "{sess.sport}")'
                cell.font = link_font
            else:
                ws.cell(row=row_idx, column=3, value=sess.sport)

            # Venue as link
            if sess.venue in all_venues:
                cell = ws.cell(row=row_idx, column=4)
                cell.value = f'=HYPERLINK("#Venues!A{all_venues[sess.venue]}", "{sess.venue}")'
                cell.font = link_font
            else:
                ws.cell(row=row_idx, column=4, value=sess.venue)

            # Zone as link
            zone = sess.venue_rel.zone if sess.venue_rel else None
            if zone and zone in all_zones:
                cell = ws.cell(row=row_idx, column=5)
                cell.value = f'=HYPERLINK("#Zones!A{all_zones[zone]}", "{zone}")'
                cell.font = link_font
            else:
                ws.cell(row=row_idx, column=5, value=zone)

            ws.cell(row=row_idx, column=6, value=sess.type)

            # Real datetime values
            start_cell = ws.cell(row=row_idx, column=7, value=sess.starts_at.replace(tzinfo=None))
            start_cell.number_format = date_format
            end_cell = ws.cell(row=row_idx, column=8, value=sess.ends_at.replace(tzinfo=None))
            end_cell.number_format = date_format

            dur_cell = ws.cell(row=row_idx, column=9, value=sess.duration_minutes)
            dur_cell.number_format = '0'

            ws.cell(row=row_idx, column=10, value=sess.ticketed)
            ws.cell(row=row_idx, column=11, value=sess.venue_rel.in_okc if sess.venue_rel else False)

            events_cell = ws.cell(row=row_idx, column=12, value=len(sess.events))
            events_cell.number_format = '0'

        counts["sessions"] = len(sessions)

    # Add data validation for type column (dropdown)
    event_types = ["Final", "Bronze", "Semifinal", "Quarterfinal", "Repechage", "Preliminary", "N/A"]
    type_dv = DataValidation(type="list", formula1=f'"{",".join(event_types)}"', allow_blank=True)
    type_dv.error = "Select from list"
    type_dv.prompt = "Select session type"
    ws.add_data_validation(type_dv)
    type_dv.add(f"F2:F{len(sessions) + 1}")

    style_header(ws)
    add_borders(ws, len(sessions))
    auto_width(ws)

    # Events sheet
    ws_events = wb.create_sheet("Events")
    with db.session() as s:
        events = s.exec(
            select(Event)
            .options(selectinload(Event.session_rel).selectinload(Session.venue_rel))
            .order_by(Event.event_number)
        ).all()

        ws_events.append(["event_number", "code", "day", "sport", "venue", "zone",
                          "starts_at", "sex", "description", "type", "order", "total"])

        for row_idx, e in enumerate(events, 2):
            sess = e.session_rel

            evt_cell = ws_events.cell(row=row_idx, column=1, value=e.event_number)
            evt_cell.number_format = '0'

            # Session code as link
            if e.code in all_sessions:
                cell = ws_events.cell(row=row_idx, column=2)
                cell.value = f'=HYPERLINK("#Sessions!A{all_sessions[e.code]}", "{e.code}")'
                cell.font = link_font
            else:
                ws_events.cell(row=row_idx, column=2, value=e.code)

            day_cell = ws_events.cell(row=row_idx, column=3, value=sess.day if sess else None)
            day_cell.number_format = '0'

            # Sport as link
            sport = sess.sport if sess else None
            if sport and sport in all_sports:
                cell = ws_events.cell(row=row_idx, column=4)
                cell.value = f'=HYPERLINK("#Sports!A{all_sports[sport]}", "{sport}")'
                cell.font = link_font
            else:
                ws_events.cell(row=row_idx, column=4, value=sport)

            # Venue as link
            venue = sess.venue if sess else None
            if venue and venue in all_venues:
                cell = ws_events.cell(row=row_idx, column=5)
                cell.value = f'=HYPERLINK("#Venues!A{all_venues[venue]}", "{venue}")'
                cell.font = link_font
            else:
                ws_events.cell(row=row_idx, column=5, value=venue)

            # Zone as link
            zone = sess.venue_rel.zone if sess and sess.venue_rel else None
            if zone and zone in all_zones:
                cell = ws_events.cell(row=row_idx, column=6)
                cell.value = f'=HYPERLINK("#Zones!A{all_zones[zone]}", "{zone}")'
                cell.font = link_font
            else:
                ws_events.cell(row=row_idx, column=6, value=zone)

            if sess:
                start_cell = ws_events.cell(row=row_idx, column=7, value=sess.starts_at.replace(tzinfo=None))
                start_cell.number_format = date_format

            ws_events.cell(row=row_idx, column=8, value=e.sex)
            ws_events.cell(row=row_idx, column=9, value=e.description)
            ws_events.cell(row=row_idx, column=10, value=e.type)

            order_cell = ws_events.cell(row=row_idx, column=11, value=e.order_in_session)
            order_cell.number_format = '0'
            total_cell = ws_events.cell(row=row_idx, column=12, value=e.total_in_session)
            total_cell.number_format = '0'

        counts["events"] = len(events)

    # Sex dropdown
    sex_dv = DataValidation(type="list", formula1='"Men,Women,Mixed,and/or,or"', allow_blank=True)
    ws_events.add_data_validation(sex_dv)
    sex_dv.add(f"H2:H{len(events) + 1}")

    # Type dropdown
    type_dv2 = DataValidation(type="list", formula1=f'"{",".join(event_types)}"', allow_blank=True)
    ws_events.add_data_validation(type_dv2)
    type_dv2.add(f"J2:J{len(events) + 1}")

    style_header(ws_events)
    add_borders(ws_events, len(events))
    auto_width(ws_events)

    # Reorder sheets: Sessions, Events, Sports, Venues, Zones
    wb._sheets = [wb["Sessions"], wb["Events"], wb["Sports"], wb["Venues"], wb["Zones"]]

    # Remove default empty sheet if exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(path)
    return counts


def export_all(db: Database, output_dir: str | Path) -> dict[str, int]:
    """Export all data to JSON and CSV files."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    return {
        "sessions_json": export_sessions_json(db, output_dir / "sessions.json"),
        "sessions_csv": export_sessions_csv(db, output_dir / "sessions.csv"),
        "events_json": export_events_json(db, output_dir / "events.json"),
        "events_csv": export_events_csv(db, output_dir / "events.csv"),
        "sports_json": export_sports_json(db, output_dir / "sports.json"),
        "sports_csv": export_sports_csv(db, output_dir / "sports.csv"),
        "venues_json": export_venues_json(db, output_dir / "venues.json"),
        "venues_csv": export_venues_csv(db, output_dir / "venues.csv"),
        "zones_json": export_zones_json(db, output_dir / "zones.json"),
        "zones_csv": export_zones_csv(db, output_dir / "zones.csv"),
        "la28.xlsx": export_xlsx(db, output_dir / "la28.xlsx"),
    }
